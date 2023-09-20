import openpyxl


#Inicializa el driver cuando se llama la función
class Funexcel():
    def __init__(self, driver):
        self.driver = driver


# cuenta el numero de filas que tiene la hoja de excel,
# pide un path (ruta) , la sheetName (nombre de hoja)
#cuenta el maximo de filas
    def getRowCount(file,path,sheetName):
        Worbook=openpyxl.load_workbook(path)
        sheet= Worbook[sheetName]
        return (sheet.max_row)

# cuenta el numero de columnas que tiene la hoja de excel,
# pide la sheetName (nombre de hoja)
# cuenta el maximo de filas
    def getColumnCount(file,sheetName):
        Worbook= openpyxl.load_workbook(file)
        sheet=  Worbook[sheetName]
        return (sheet.max_column)

#Lee los datos, pide la dirección, el nombre de hoja, el numero de columnas y el de filas
    def readData(file,path,sheetName,rownum,columnno):
        Worbook = openpyxl.load_workbook(path)
        sheet =  Worbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

# Toma el valor "data" y lo guarda/imprime
    def writeData(file,path,sheetName,rownum,columnno,data):
        Worbook = openpyxl.load_workbook(path)
        sheet =  Worbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        Worbook.save(path)



